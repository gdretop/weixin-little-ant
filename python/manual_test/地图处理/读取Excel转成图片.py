#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import traceback

import openpyxl
from PIL import Image
from PIL import ImageDraw
from PIL.ImageEnhance import Color
from openpyxl.styles import PatternFill

from config.mori_map_config import SCALE_SIZE
import common.image_tool
from game_map.item_map import type_map

"""
位置是Excel上的位置
0  位置 01  01  00B050 [  0, 176,  80] 深绿  草地岩石-fgColor.rgb
1  位置 94  26  92D050 [146, 208,  80] 浅绿  栏栅-fgColor.rgb
2  位置 07  05  F0F0F0 [  0,   0,   0] 黑色  普通地块-fgColor.rgb
3  位置 06  05  7030A0 [112,  48, 160] 紫色  山洞-fgColor.rgb  
4  位置 12  17  F196AB [241, 150, 171] 粉红  坐标牌-fgColor.rgb
5  位置 94  25  FF922B [255, 146,  43] 橘色  建筑-fgColor.theme=5
6  位置 149 152 00B0F0 [  0, 176, 240] 浅蓝  商店-fgColor.rgb
7  位置 151 148 817709 [129, 119,   9] 橙色  医院-fgColor.rgb
8  位置 140 95  FFC000 [255, 192,   0] 浅黄  教堂-fgColor.rgb
9  位置 149 148 FF0000 [255,   0,   0] 大红  加油站-fgColor.rgb
10 位置 144 124 002060 [  0,  32,  96] 深蓝  广场-fgColor.rgb
11 位置 132 94  C00000 [240, 240, 240] 深红  帐篷-fgColor.rgb 
12 位置 162 32  4290FF [ 66, 144, 255] 青蓝  工厂-fgColor.rgb 
13 位置 170 28  A219FF [162,  25, 255] 浅紫  房屋-fgColor.rgb 
14 位置 261 91  598F38 [89, 143, 56] 深绿  热带树-fgColor.rgb 
15 位置 264 87  E5F3F1 [229, 243, 241] 白灰  石头-fgColor.rgb 
16 位置 260 87  355522 [ 53,  85,  34] 深绿  深绿树-fgColor.rgb 
17 位置 274 130 BEFF9E [190, 255, 158] 浅绿  带花草-fgColor.rgb 
18 位置 263 99  89FF8B [137, 255, 139] 浅绿  小草-fgColor.rgb 
19 位置 85 185  FAFF9F [250, 255, 159] 浅黄  标记空地-fgColor.rgb 
"""
type_map_excel = {
    '00B050': 0,
    '92D050': 1,
    'F0F0F0': 2,
    '7030A0': 3,
    'F196AB': 4,
    'FF922B': 5,
    '00B0F0': 6,
    '817709': 7,
    'FFC000': 8,
    'FF0000': 9,
    '002060': 10,
    'C00000': 11,
    '4290FF': 12,
    'A219FF': 13,
    '598F38': 14,
    'E5F3F1': 15,
    '355522': 16,
    'BEFF9E': 17,
    '89FF8B': 18,
    'FAFF9F': 19,
}
# https://www.cnblogs.com/zxt518/p/15430700.html excel 操作
dir = '/Users/yuwanglin/project/weixin-little-ant/python/manual_test/地图处理/'
file_path_1 = dir + '生存之路6.0.1.1.xlsx'
file_path_2 = dir + '末日生存第二张地图.xlsx'
file_path_3 = dir + '生存之路扩张.xlsx'
file_path_4 = dir + '合并地图.xlsx'
file_path_5 = dir + '飞行猫的文档.xlsx'
ele_dir = dir + '元素/'
output_path = dir + 'mori_game_map.png'
output_gray_path = dir + 'mori_game_map_gray.png'
output_whole_path = dir + 'whole_image.png'


def trans_excel_2_png(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb[sheet_name]
    # print("Sheet2:{0}*{1}的表格".format(sheet.max_row, sheet.max_column))
    image = Image.new('RGB', (301, 301), (255, 255, 255))  # 分别是颜色模式、尺寸、颜色
    image_gray = Image.new('L', (301, 301))
    im = image.load()
    im_gray = image_gray.load()
    for row_index, row in enumerate(sheet.rows):
        for column_index, cell in enumerate(row):
            # if row_index == 4:
            #     print(row_index)
            try:
                value = ''
                if cell.fill and cell.fill.fgColor:
                    if cell.fill.fgColor.theme == 5:
                        value = 'FF922B'
                    elif cell.fill.fgColor.rgb:
                        value = cell.fill.fgColor.rgb[2:8]
                    else:
                        print("问题坐标{},{}".format(row_index, column_index))
                    if value == '000000':
                        value = 'F0F0F0'
                else:
                    print("问题坐标{},{}".format(row_index, column_index))
                if value:
                    # im[column_index, row_index] = common.image_tool.str_2_rgb(value)
                    im[column_index, row_index] = common.image_tool.str_2_rgb(type_map[type_map_excel[value]]['target_color'])
                    im_gray[column_index, row_index] = type_map_excel[value] * SCALE_SIZE
            except Exception as e:
                print(e)
    # image.show()  # 展示，可不要
    scale_size = 4
    new_size = 300 * scale_size
    image_big = Image.new('RGB', (new_size, new_size), (255, 255, 255))  # 分别是颜色模式、尺寸、颜色
    image_big_im = image_big.load()
    for i in range(new_size):
        for j in range(new_size):
            try:
                image_big_im[i, j] = im[int(i / scale_size), int(j / scale_size)]
            except BaseException as e:
                print(e)
    image_big = image_big.rotate(-45, expand=True)
    image_big.save(output_path)  # 保存路径及名称
    image_gray.save(output_gray_path)
    return image_gray, im_gray


# trans_excel_2_png()

"""
树
山
路牌
教堂
油站
广场
建筑
山洞
医院
商店
帐篷
工厂
栏
"""
type_map_2 = {
    "树": [0, [0, 14, 16, 17, 18]],
    "栏": [1, [1]],
    "空地": [2, [2]],
    "山洞": [3, [3]],
    "路牌": [4, [4]],
    "建筑": [5, [5, 13]],
    "商店": [6, [6]],
    "医院": [7, [7]],
    "教堂": [8, [8]],
    "油站": [9, [9]],
    "广场": [10, [10]],
    "帐篷": [11, [11]],
    "工厂": [12, [12]],
    "房屋": [13, [13]],
    "热带树": [14, [0, 14]],
    "山": [15, [0, 15]],
    "深绿树": [16, [16]],
    "带花草": [17, [17]],
    "小草": [18, [18]],
}

second_type_map = {
    "树": 16,
    "山": 15,
    "路牌": 4,
    "教堂": 8,
    "油站": 9,
    "广场": 10,
    "建筑": 13,
    "山洞": 3,
    "医院": 7,
    "商店": 6,
    "帐篷": 11,
    "工厂": 12,
    "栏": 1,
}


def compare_2_excel():
    image_gray, image = trans_excel_2_png(file_path_3, 'scale_map')
    wb = openpyxl.load_workbook(file_path_2, read_only=True)
    sheet = wb['Sheet2']
    for row_index, row in enumerate(sheet.rows):
        # if row_index < 101 or row_index > 111:
        #     continue
        if row_index >= 300:
            break
        for column_index, cell in enumerate(row):
            # if column_index < 25 or column_index > 35:
            #     continue
            if column_index >= 300:
                break
            try:
                value = cell.value
                if value not in type_map_2:
                    value = "空地"
                info = type_map_2[value]
                type = int(image[column_index, row_index] / SCALE_SIZE)
                if not type_map[type]['reach'] or type >= 6:
                    continue
                # if type not in info[1]:
                # if value in ['广场','帐篷','工厂','教堂','山洞','医院','油站','商店'] and type_map[type]['name'] in ['建筑']:
                #     continue
                if value not in ['栏', '山', '树']:
                    continue
                    # if type not in
                print("找到不同位置:", column_index + 1, row_index + 1, value, type_map[type]['name'])
                image[column_index, row_index] = 0
            except Exception as e:
                traceback.print_exc()
                print(e)
    image_gray.save(output_gray_path)


# compare_2_excel()
def map_scale():
    image_gray, image = trans_excel_2_png(file_path_1, 'Sheet1')
    wb = openpyxl.load_workbook(file_path_1, read_only=False)
    sheet = wb.copy_worksheet(wb.worksheets[1])
    sheet.title = 'scale_map'
    wb.active = sheet
    for row_index, row in enumerate(sheet.rows):
        if row_index >= 300:
            break
        for column_index, cell in enumerate(row):
            if column_index >= 300:
                break
            if image[column_index, row_index] != 2 * SCALE_SIZE:
                continue
            flag = False
            for i in range(-3, 4, 1):
                for j in range(-3, 4, 1):
                    if abs(i) + abs(j) > 3:
                        continue
                    r_i = row_index + i
                    c_j = column_index + j
                    if r_i < 0 or r_i >= 300 or c_j < 0 or c_j >= 300:
                        continue
                    if image[c_j, r_i] == 19 * SCALE_SIZE:
                        flag = True
            if flag:
                cell.fill = PatternFill('solid', fgColor='FAFF9F')
    wb.save(file_path_3)


def combine_map_2_excel():
    image_gray, image = trans_excel_2_png(file_path_1, 'Sheet1')
    wb = openpyxl.load_workbook(file_path_1, read_only=False)
    sheet = wb.copy_worksheet(wb.worksheets[1])
    sheet.title = 'combine_map'
    wb.active = sheet

    wb2 = openpyxl.load_workbook(file_path_2, read_only=False)
    sheet2 = wb2['Sheet2']
    for row_index, row in enumerate(sheet.rows):
        if row_index >= 300:
            break
        row2 = sheet2[row_index + 1]
        for column_index, cell in enumerate(row):
            if column_index >= 300:
                break
            # 已标记的位置不再更新
            if image[column_index, row_index] >= 6 * SCALE_SIZE:
                continue
            cell2 = row2[column_index]
            value2 = cell2.value
            # 第二张图非物体不处理
            if value2 not in second_type_map:
                continue
            # 在走过点可视区域的不合并
            flag = False
            for i in range(-3, 4, 1):
                for j in range(-3, 4, 1):
                    if abs(i) + abs(j) > 3:
                        continue
                    r_i = row_index + i
                    c_j = column_index + j
                    if r_i < 0 or r_i >= 300 or c_j < 0 or c_j >= 300:
                        continue
                    if image[c_j, r_i] == 19 * SCALE_SIZE:
                        flag = True

            if flag:
                continue
            try:
                cover_type = second_type_map[value2]
                cover_cell_info = type_map[cover_type]
                cell.fill = PatternFill('solid', fgColor=cover_cell_info['color'])
                cell.value = '?'
            except Exception as e:
                traceback.print_exc()
                print(e)
    wb.save(file_path_3)


def build_whole_map():
    ele_size = 32
    width = 300
    image_size = width * ele_size
    ele_map = {}
    ele_map_init = {}
    for i in range(20):
        path = ele_dir + "/{}.png".format(i)
        image = Image.open(path)
        image = image.resize((ele_size, ele_size))
        image.save("/Users/yuwanglin/project/weixin-little-ant/python/game_map/image/{}.png".format(i))
        ele_map_init[i] = image
        image = image.load()
        ele_map[i] = image
    image = Image.new('RGB', (image_size, image_size), (255, 255, 255))  # 分别是颜色模式、尺寸、颜色
    im = image.load()
    image_gray, im_gray = trans_excel_2_png(file_path_4, '颜色地图')
    i_offset = 0
    j_offset = 0
    for i in range(width):
        for j in range(width):
            value = int(im_gray[i+i_offset, j+j_offset] / SCALE_SIZE)
            if value not in [2,19]:
                for k in range(ele_size):
                    for l in range(ele_size):
                        im[i * ele_size + k, j * ele_size + l] = ele_map[value][k, l]
            else:
                copy_img = ele_map_init[value].copy()
                draw = ImageDraw.Draw(copy_img,'RGBA')
                draw.text((0,0),"x {}\ny {}".format(i+1+i_offset,j+1+j_offset),fill=(255,255,255))
                im_text = copy_img.load()
                for k in range(ele_size):
                    for l in range(ele_size):
                        im[i * ele_size + k, j * ele_size + l] = im_text[k, l]

    image.save(output_whole_path)


trans_excel_2_png(file_path_4, '颜色地图')
build_whole_map()
# map_scale()
# compare_2_excel()
# print([i for i in range(-3, 4, 1)])
# 合并两张图
# combine_map_2_excel()
# print("从当前点92,198出发,需要走311步,可经过建筑26\n第1次:X+1步到达(93,198)\n第2次:Y-2步到达(93,196)\n第3次:X+34步到达(127,196)\n路上经过['帐篷', '房屋']\n第4次:Y-15步到达(127,181)\n路上经过['加油站']\n第5次:X+2步到达(129,181)\n第6次:Y-3步到达(129,178)\n第7次:X+1步到达(130,178)\n第8次:Y-4步到达(130,174)\n第9次:X+10步到达(140,174)\n路上经过['坐标牌']\n第10次:Y-6步到达(140,168)\n路上经过['房屋']\n第11次:X+10步到达(150,168)\n第12次:Y-6步到达(150,162)\n路上经过['广场']\n第13次:X+21步到达(171,162)\n第100步你将【停留在156,162】\n路上经过['房屋']\n第14次:Y-2步到达(171,160)\n路上经过['房屋']\n第15次:X+4步到达(175,160)\n第16次:Y-7步到达(175,153)\n路上经过['医院']\n第17次:X+9步到达(184,153)\n第18次:Y-1步到达(184,152)\n第19次:X+8步到达(192,152)\n路上经过['房屋']\n第20次:Y-4步到达(192,148)\n路上经过['房屋']\n第21次:X+5步到达(197,148)\n第22次:Y-2步到达(197,146)\n第23次:X+17步到达(214,146)\n路上经过['广场', '房屋']\n第24次:Y-5步到达(214,141)\n第25次:X+4步到达(218,141)\n路上经过['帐篷']\n第26次:Y-8步到达(218,133)\n第27次:X+2步到达(220,133)\n第28次:Y-4步到达(220,129)\n第29次:X+3步到达(223,129)\n路上经过['房屋']\n第30次:Y-4步到达(223,125)\n第31次:X+3步到达(226,125)\n路上经过['加油站']\n第32次:Y-12步到达(226,113)\n路上经过['房屋']\n第33次:X+6步到达(232,113)\n路上经过['工厂']\n第34次:Y-6步到达(232,107)\n第35次:X+6步到达(238,107)\n路上经过['坐标牌']\n第36次:Y-3步到达(238,104)\n路上经过['房屋']\n第37次:X+2步到达(240,104)\n第38次:Y-3步到达(240,101)\n第39次:X+12步到达(252,101)\n路上经过['商店']\n第40次:Y-3步到达(252,98)\n第41次:X+11步到达(263,98)\n第42次:Y-12步到达(263,86)\n路上经过['房屋']\n第43次:X+2步到达(265,86)\n第44次:Y-3步到达(265,83)\n路上经过['工厂']\n第45次:X+4步到达(269,83)\n路上经过['房屋']\n第46次:Y-8步到达(269,75)\n路上经过['帐篷']\n第47次:X+9步到达(278,75)\n路上经过['房屋']\n第48次:Y-2步到达(278,73)\n\n公众号:旺仔小蚂蚁")