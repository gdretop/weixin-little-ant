#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import traceback

import openpyxl
from PIL import Image
from config.mori_map_config import SCALE_SIZE
import common.image_tool
from game_map.item_map import type_map

"""
位置是Excel上的位置
0  位置 01  01  00B050 [  0, 176,  80] 深绿  草地岩石-fgColor.rgb
1  位置 94  26  92D050 [146, 208,  80] 浅绿  栏栅-fgColor.rgb
2  位置 07  05  F0F0F0 [  0,   0,   0] 黑色  普通地块-fgColor.rgb
3  位置 06  05  7030A0 [112,  48, 160] 紫色  山洞-fgColor.rgb  
4  位置 12  17  F196AB [241, 150, 171] 粉红  坐标牌-fgColor.rgb
5  位置 94  25  FF922B [255, 146,  43] 橘色  建筑-fgColor.theme=5
6  位置 149 152 00B0F0 [  0, 176, 240] 浅蓝  商店-fgColor.rgb
7  位置 151 148 817709 [129, 119,   9] 橙色  医院-fgColor.rgb
8  位置 140 95  FFC000 [255, 192,   0] 浅黄  教堂-fgColor.rgb
9  位置 149 148 FF0000 [255,   0,   0] 大红  加油站-fgColor.rgb
10 位置 144 124 002060 [  0,  32,  96] 深蓝  广场-fgColor.rgb
11 位置 132 94  C00000 [240, 240, 240] 深红  帐篷-fgColor.rgb 
12 位置 162 32  4290FF [ 66, 144, 255] 青蓝  工厂-fgColor.rgb 
13 位置 170 28  A219FF [162,  25, 255] 浅紫  房屋-fgColor.rgb 
14 位置 261 91  598F38 [89, 143, 56] 深绿  热带树-fgColor.rgb 
15 位置 264 87  E5F3F1 [229, 243, 241] 白灰  石头-fgColor.rgb 
16 位置 260 87  355522 [ 53,  85,  34] 深绿  深绿树-fgColor.rgb 
17 位置 274 130 BEFF9E [190, 255, 158] 浅绿  带花草-fgColor.rgb 
18 位置 263 99  89FF8B [137, 255, 139] 浅绿  小草-fgColor.rgb 
"""
type_map_excel = {
    "栏": 1,
    "空地":2,
    "山洞": 3,
    "路牌": 4,
    "建筑": 5,
    "商店": 6,
    "医院": 7,
    "教堂": 8,
    "油站": 9,
    "广场": 10,
    "帐篷": 11,
    "工厂": 12,
    "山": 15,
    "树": 16,
}
# https://www.cnblogs.com/zxt518/p/15430700.html excel 操作
dir = '/Users/yuwanglin/project/weixin-little-ant/python/manual_test/僵尸防线/'
file_path = dir + '末日生存第二张地图.xlsx'
file_path_2 = dir + '末日生存第二张地图.xlsx'
output_path = dir + 'jiangshi_game_map.png'
output_gray_path = dir + 'jiangshi_game_map_gray.png'


def trans_excel_2_png():
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb['Sheet2']
    # print("Sheet2:{0}*{1}的表格".format(sheet.max_row, sheet.max_column))
    # image = Image.new('RGB', (301, 301), (255, 255, 255))  # 分别是颜色模式、尺寸、颜色
    image_gray = Image.new('L', (301, 301))
    # im = image.load()
    im_gray = image_gray.load()
    for row_index, row in enumerate(sheet.rows):
        for column_index, cell in enumerate(row):
            try:
                value = cell.value
                # if value is None or value in type_map_excel:
                #     continue
                # print(value)
                if value is None or value not in type_map_excel:
                    value = '空地'
                # im[column_index, row_index] = common.image_tool.str_2_rgb(value)
                im_gray[column_index, row_index] = type_map_excel[value] * SCALE_SIZE
            except Exception as e:
                print(e)
    # image.show()  # 展示，可不要
    # image.save(output_path)  # 保存路径及名称
    image_gray.save(output_gray_path)
    return im_gray


trans_excel_2_png()
